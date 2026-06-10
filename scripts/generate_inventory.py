import os
import yaml
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

TEST_CASES_DIR = "Test Suites"
CSV_FILE = "OnTrack TCM - Test Suites.csv"
XLSX_FILE = "OnTrack TCM - Test Suites.xlsx"

FIELDS = ["Test Cases", "Test Schedule", "Assigned Tester",
          "Date of Execution", "Status", "Bug ID"]

def extract_frontmatter(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    if content.startswith("---"):
        parts = content.split("---", 2)
        frontmatter = yaml.safe_load(parts[1]) or {}
        return frontmatter
    return {}

def collect_testcases():
    sections = {}
    for root, _, files in os.walk(TEST_CASES_DIR):
        section_name = os.path.basename(root)
        if not files:
            continue
        rows = []
        for file in files:
            if file.endswith(".md"):
                path = os.path.join(root, file)
                fm = extract_frontmatter(path)
                rows.append({
                   "Test Cases": fm.get("Title", file.replace(".md", "")),
                    "Test Schedule": fm.get("Test Schedule", "TBD"),
                    "Assigned Tester": fm.get("Assigned Tester", "TBD"),
                    "Date of Execution": fm.get("Date of Execution", "TBD"),
                    "Status": fm.get("Status", "Draft"),
                    "Bug ID": fm.get("Bug ID", "")
                })
        if rows:
            sections[section_name] = rows
    return sections

def write_csv(sections):
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=FIELDS)
        writer.writeheader()

        for section, rows in sections.items():
            writer.writerow({})
            writer.writerow({"Test Cases": section})
            for row in rows:
                writer.writerow(row)

def write_xlsx(sections, filename="OnTrack TCM - Test Suites.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Suites"

    # Header row
    ws.append(FIELDS)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for col in range(1, len(FIELDS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Section styling
    section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Write sections
    for section, rows in sections.items():
        ws.append([])
        ws.append([section])
        section_cell = ws.cell(row=ws.max_row, column=1)
        section_cell.font = Font(bold=True)
        section_cell.fill = section_fill
        for row in rows:
            ws.append([row[field] for field in FIELDS])

    # Adjust column widths
    for i, field in enumerate(FIELDS, 1):
        max_length = max(len(str(ws.cell(row=r, column=i).value or "")) for r in range(1, ws.max_row+1))
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Conditional formatting for Status column
    status_col = FIELDS.index("Status") + 1
    ws.conditional_formatting.add(
        f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{ws.max_row}",
        CellIsRule(operator="equal", formula=['"PASSED"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{ws.max_row}",
        CellIsRule(operator="equal", formula=['"FAILED"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{ws.max_row}",
        CellIsRule(operator="equal", formula=['"Draft"'],
                   fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"))
    )

    # Dropdown for Status column
    dv = DataValidation(type="list", formula1='"PASSED,FAILED,Draft"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{ws.max_row}")

    # Date format for Test Schedule & Date of Execution
    for col_name in ["Test Schedule", "Date of Execution"]:
        col_idx = FIELDS.index(col_name) + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = "yyyy/mm/dd"

    wb.save(filename)
    
def main():
    sections = collect_testcases()
    write_csv(sections)
    write_xlsx(sections)
    print(f"✅ Inventory generated: {CSV_FILE} and {XLSX_FILE}")

if __name__ == "__main__":
    main()
